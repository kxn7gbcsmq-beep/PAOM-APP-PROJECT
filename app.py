import streamlit as st
from openpyxl import load_workbook
import io

st.title("Анализ оценок по Excel")

uploaded_files = st.file_uploader(
    "Загрузите Excel-файл(ы) с оценками",
    accept_multiple_files=True,
    type=["xlsx", "xlsm"]
)

for uploaded_file in uploaded_files:
    st.subheader(f"Файл: {uploaded_file.name}")
    try:
        in_memory_file = io.BytesIO(uploaded_file.read())
        book = load_workbook(in_memory_file, data_only=True)
        sheet = book.active
        
        subjects = []
        data = {}

        for cell in sheet[1]:
            if cell.value is not None:
                subjects.append(cell.value)
                data[cell.value] = []

        for row in sheet.iter_rows(min_row=1, values_only=True):
            for i, value in enumerate(row):
                if i >= len(subjects):
                    continue
                if value is None:
                    continue
                try:
                    value = int(value)
                except:
                    continue
                if 2 <= value <= 5:
                    data[subjects[i]].append(value)
                else:
                    st.error(f'Ошибка! Найдена цифра, не являющаяся оценкой: {value}')
                    continue

        for subject, grades in data.items():
            if not grades:
                st.write(f'\nНет оценок по предмету: {subject}')
                continue

            st.markdown('---')
            st.write(f'Анализ оценок по предмету: {subject}')

            count = len(grades)
            average = sum(grades) / count
            minimum = min(grades)
            maximum = max(grades)
            sorted_grades = sorted(grades)
            if count % 2 == 1:
                median = sorted_grades[count // 2]
            else:
                median = (sorted_grades[count // 2 - 1] + sorted_grades[count // 2]) / 2
            mode = max(set(grades), key=grades.count)

            st.write("Количество оценок:", count)
            st.write("Средний балл:", round(average, 2))
            st.write("Итоговая оценка:", round(average))
            if 2.3 < average < 2.5:
                st.write("Не хватает несколько положительных оценок до итоговой оценки 3")
            if 3.3 < average < 3.5:
                st.write("Не хватает несколько положительных оценок до итоговой оценки 4")
            if 4.3 < average < 4.5:
                st.write("Не хватает несколько положительных оценок до итоговой оценки 5")
            st.write("Минимальная оценка:", minimum)
            st.write("Максимальная оценка:", maximum)
            st.write("Медиана:", median)
            st.write("Мода:", mode)

    except Exception as e:
        st.error(f"Не удалось прочитать Excel-файл: {e}")
